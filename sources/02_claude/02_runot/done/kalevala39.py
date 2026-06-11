from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Runo 39"

# Headers
headers = ["A: Alkuperäinen säe", "B: Nykysuomi", "C: Perusmuodot", "D: Käännös (EN)", "E: Trickster-huomiot"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, name="Arial", size=11, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="2E4057")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data: (original, modern, basic_forms, translation, trickster)
verses = [
    (
        "Vaka vanha Väinämöinen",
        "Vankka vanha Väinämöinen",
        "vakaa, vanha, Väinämöinen",
        "steady, old, Väinämöinen",
        ""
    ),
    (
        "itse tuon sanoiksi virkki:",
        "itse tuon sanoiksi virkki:",
        "itse, tuo, sana, virkkoa",
        "himself, that, word/speech, to say/utter",
        ""
    ),
    (
        '"Ohoh seppo Ilmarinen!',
        '"Ohoh seppä Ilmarinen!',
        "ohoh, seppä, Ilmarinen",
        "oh/hey, smith, Ilmarinen",
        ""
    ),
    (
        "Lähtekämme Pohjolahan",
        "Lähdetäämme Pohjolaan",
        "lähteä, Pohjola",
        "to leave/depart, Pohjola (the North)",
        ""
    ),
    (
        "hyvän sammon saa'antahan,",
        "hyvän sammon saantiin,",
        "hyvä, sampo, saanti",
        "good, Sampo (magic mill), getting/obtaining",
        ""
    ),
    (
        "kirjokannen katsantahan!",
        "kirjokannen katsomiseen!",
        "kirjokansi, katsanto",
        "decorated lid, viewing/looking at",
        ""
    ),
    (
        "Se on seppo Ilmarinen",
        "Se on seppä Ilmarinen",
        "se, seppä, Ilmarinen",
        "he/it, smith, Ilmarinen",
        ""
    ),
    (
        "sanan virkkoi, noin nimesi:",
        "sanan sanoi, noin nimesi:",
        "sana, virkkoa, noin, nimetä",
        "word, to say, thus/so, to name/call",
        ""
    ),
    (
        '"Ei ole sampo saatavana,',
        '"Ei ole sampo saatavissa,',
        "ei, olla, sampo, saatava",
        "no/not, to be, Sampo, obtainable/gettable",
        ""
    ),
    (
        "kirjokansi tuotavana",
        "kirjokansi tuotavissa",
        "kirjokansi, tuotava",
        "decorated lid, bringable/to be brought",
        ""
    ),
    (
        "pimeästä Pohjolasta,",
        "pimeästä Pohjolasta,",
        "pimeä, Pohjola",
        "dark, Pohjola (the North)",
        ""
    ),
    (
        "summasta Sariolasta!",
        "hämärästä Sariolasta!",
        "summa, Sariola",
        "dim/murky, Sariola (another name for Pohjola)",
        ""
    ),
    (
        "Siell' on sampo saatettuna,",
        "Siellä on sampo saatettuna,",
        "siellä, sampo, saattaa",
        "there, Sampo, to escort/bring/place",
        ""
    ),
    (
        "kirjokansi kannettuna",
        "kirjokansi kannettuna",
        "kirjokansi, kantaa",
        "decorated lid, to carry/bear",
        ""
    ),
    (
        "Pohjolan kivimäkehen,",
        "Pohjolan kivimäkeen,",
        "Pohjola, kivimäki",
        "Pohjola, rocky/stone hill",
        ""
    ),
    (
        "vaaran vaskisen sisähän",
        "vaaran vaskisen sisään",
        "vaara, vaskinen, sisä",
        "hill/fell, copper/bronze (adj.), inside/interior",
        ""
    ),
    (
        "yheksän lukon ta'aksi;",
        "yhdeksän lukon taakse;",
        "yhdeksän, lukko, taakse",
        "nine, lock, behind",
        ""
    ),
    (
        "siihen juuret juurruteltu",
        "siihen juuret juurruteltu",
        "se, juuri, juurruttaa",
        "it/that, root, to root/anchor",
        ""
    ),
    (
        "yheksän sylen syvähän,",
        "yhdeksän sylen syvyyteen,",
        "yhdeksän, syli, syvyys",
        "nine, fathom (arm-span), depth",
        ""
    ),
    (
        "yksi juuri maaemähän,",
        "yksi juuri maaemään,",
        "yksi, juuri, maaemo",
        "one, root, Mother Earth",
        ""
    ),
    (
        "toinen vesiviertehesen,",
        "toinen vesivirteen,",
        "toinen, vesivirta",
        "second/another, water current/stream",
        ""
    ),
    (
        "kolmas on kotimäkehen.",
        "kolmas on kotimäkeen.",
        "kolmas, koti, mäki",
        "third, home, hill",
        ""
    ),
    (
        "Sanoi vanha Väinämöinen:",
        "Sanoi vanha Väinämöinen:",
        "sanoa, vanha, Väinämöinen",
        "to say, old, Väinämöinen",
        ""
    ),
    (
        '"Veli seppo, veikkoseni!',
        '"Veli seppä, veikkoseni!',
        "veli, seppä, veikko",
        "brother, smith, brother/buddy",
        ""
    ),
    (
        "Lähtekämme Pohjolahan",
        "Lähdetäämme Pohjolaan",
        "lähteä, Pohjola",
        "to leave/depart, Pohjola",
        ""
    ),
    (
        "tuon on sammon saa'antahan!",
        "tuon sammon saantiin!",
        "tuo, sampo, saanti",
        "that, Sampo, getting/obtaining",
        ""
    ),
    (
        "Laatikamme laiva suuri,",
        "Laittakaamme laiva suuri,",
        "laittaa, laiva, suuri",
        "to make/prepare, ship, large/great",
        ""
    ),
    (
        "johon sampo saatetahan,",
        "johon sampo saadaan,",
        "joka, sampo, saattaa",
        "which/that, Sampo, to bring/convey",
        ""
    ),
    (
        "kirjokansi kannetahan",
        "kirjokansi kannetaan",
        "kirjokansi, kantaa",
        "decorated lid, to carry",
        ""
    ),
    (
        "Pohjolan kivimäestä,",
        "Pohjolan kivimäestä,",
        "Pohjola, kivimäki",
        "Pohjola, rocky hill",
        ""
    ),
    (
        "vaaran vaskisen sisästä,",
        "vaaran vaskisen sisästä,",
        "vaara, vaskinen, sisä",
        "hill, copper/bronze (adj.), inside",
        ""
    ),
    (
        "yheksän lukon takoa!",
        "yhdeksän lukon takaa!",
        "yhdeksän, lukko, takaa",
        "nine, lock, from behind",
        ""
    ),
    (
        "Sanoi seppo Ilmarinen:",
        "Sanoi seppä Ilmarinen:",
        "sanoa, seppä, Ilmarinen",
        "to say, smith, Ilmarinen",
        ""
    ),
    (
        '"Vakavampi maisin matka.',
        '"Vakavampi mainen matka.',
        "vakava, mainen, matka",
        "serious/steady, land (adj.), journey",
        ""
    ),
    (
        "Lempo menköhön merelle,",
        "Lempo menköön merelle,",
        "lempo, mennä, meri",
        "devil/fiend, to go, sea",
        ""
    ),
    (
        "surma suurelle selälle!",
        "surma suurelle selälle!",
        "surma, suuri, selkä",
        "death/doom, large/great, open water/back",
        ""
    ),
    (
        "Siellä tuuli turjuttaisi,",
        "Siellä tuuli turjuttaisi,",
        "siellä, tuuli, turjuttaa",
        "there, wind, to batter/toss",
        ""
    ),
    (
        "siellä viskaisi vihuri,",
        "siellä viskaisi vihuri,",
        "siellä, viskata, vihuri",
        "there, to toss/fling, gale/squall",
        ""
    ),
    (
        "saisi sormet soutimeksi,",
        "saisi sormet soutimeksi,",
        "saada, sormi, soudin",
        "to get/have, finger, oar",
        ""
    ),
    (
        "kämmenet käsimeloiksi.",
        "kämmenet käsimeloiksi.",
        "kämmen, käsimela",
        "palm (of hand), hand-paddle",
        ""
    ),
    (
        "Sanoi vanha Väinämöinen:",
        "Sanoi vanha Väinämöinen:",
        "sanoa, vanha, Väinämöinen",
        "to say, old, Väinämöinen",
        ""
    ),
    (
        '"Vakavampi maisin matka,',
        '"Vakavampi mainen matka,',
        "vakava, mainen, matka",
        "serious/steady, land (adj.), journey",
        ""
    ),
    (
        "vakavampi, vaikeampi,",
        "vakavampi, vaikeampi,",
        "vakava, vaikea",
        "more serious/steady, more difficult",
        ""
    ),
    (
        "vielä muuten mutkaisempi.",
        "vielä muuten mutkaisempi.",
        "vielä, muuten, mutkainen",
        "still/yet, otherwise/besides, winding/complicated",
        ""
    ),
    (
        "Lysti on venon vesillä,",
        "Hauskaa on veneen vesillä,",
        "lysti, vene, vesi",
        "fun/pleasure, boat, water",
        ""
    ),
    (
        "purren juosta jolkutella,",
        "purren juosta jolkutella,",
        "pursi, juosta, jolkutella",
        "boat, to run, to glide/trot along",
        ""
    ),
    (
        "ve'et väljät välkytellä,",
        "vedet väljät välkytellä,",
        "vesi, väljä, välkyttää",
        "water, wide/open, to glitter/shimmer",
        ""
    ),
    (
        "selät selvät seurustella:",
        "selät selvät seurustella:",
        "selkä, selvä, seurustella",
        "open water/back, clear/open, to keep company/travel",
        ""
    ),
    (
        "tuuli purtta tuuittavi,",
        "tuuli purtta tuudittaa,",
        "tuuli, pursi, tuudittaa",
        "wind, boat, to rock/lull",
        ""
    ),
    (
        "aalto laivoa ajavi,",
        "aalto laivoa ajaa,",
        "aalto, laiva, ajaa",
        "wave, ship, to drive/sail",
        ""
    ),
    (
        "länsituuli läikyttävi,",
        "länsituuli läikyttää,",
        "länsituuli, läikyttää",
        "west wind, to glitter/flash",
        ""
    ),
    (
        "etelä e'elle viepi.",
        "etelä eteenpäin vie.",
        "etelä, eteenpäin, viedä",
        "south (wind), forward/onward, to carry/take",
        ""
    ),
    (
        "Vaan kuitenki kaikitenki,",
        "Vaan kuitenkin kaikitenkin,",
        "vaan, kuitenkin, kaikki",
        "but/yet, however, all/everything",
        ""
    ),
    (
        "kun et mieline merisin,",
        "kun et halua merelle,",
        "kun, ei, mieli, meri",
        "when/if, not, to desire/want, sea",
        ""
    ),
    (
        "niin on maisin matkatkamme,",
        "niin matkatkaamme maata,",
        "niin, mainen, matkustaa",
        "then/so, land (adj.), to travel",
        ""
    ),
    (
        "rantaisin ratustelkamme!",
        "rantoja pitkin kulkekaamme!",
        "ranta, ratustella",
        "shore/beach, to stroll/travel",
        ""
    ),
    (
        '"Tao nyt mulle uusi miekka,',
        '"Tao nyt minulle uusi miekka,',
        "takoa, nyt, minä, uusi, miekka",
        "to forge/hammer, now, I/me, new, sword",
        ""
    ),
    (
        "tee miekka tuliteräinen,",
        "tee miekka tuliteräinen,",
        "tehdä, miekka, tuliteräinen",
        "to make, sword, fire-steeled/razor-sharp",
        ""
    ),
    (
        "jolla hurttia hutelen,",
        "jolla hurttia hutelen,",
        "jolla, hurtta, hutella",
        "with which, dog/cur/rogue, to drive away/slash",
        ""
    ),
    (
        "Pohjan kansan kaikottelen",
        "Pohjan kansan kaikottelen",
        "pohja, kansa, kaikottaa",
        "north/bottom, people/folk, to scatter/drive off",
        ""
    ),
    (
        "saaessa otolle sammon",
        "saadessa otolle sammon",
        "saada, otto, sampo",
        "to get, taking/seizure, Sampo",
        ""
    ),
    (
        "tuonne kylmähän kylähän,",
        "tuonne kylmään kylään,",
        "tuonne, kylmä, kylä",
        "over there, cold, village",
        ""
    ),
    (
        "pimeähän Pohjolahan,",
        "pimeään Pohjolaan,",
        "pimeä, Pohjola",
        "dark, Pohjola",
        ""
    ),
    (
        "summahan Sariolahan!",
        "hämärään Sariolaan!",
        "summa, Sariola",
        "dim/murky, Sariola",
        ""
    ),
    (
        "Tuo on seppo Ilmarinen,",
        "Tuo on seppä Ilmarinen,",
        "tuo, seppä, Ilmarinen",
        "that/he, smith, Ilmarinen",
        ""
    ),
    (
        "takoja iän-ikuinen,",
        "takoja iän-ikuinen,",
        "takoja, ikä, ikuinen",
        "forger/smith, age, eternal/everlasting",
        ""
    ),
    (
        "tunki rautoja tulehen,",
        "tunki rautoja tuleen,",
        "tunkea, rauta, tuli",
        "to push/thrust, iron, fire",
        ""
    ),
    (
        "teräksiä hiiloksehen,",
        "teräksiä hiillokseen,",
        "teräs, hiillos",
        "steel, embers/coals",
        ""
    ),
    (
        "kultia koko piosen,",
        "kultia koko kourallisen,",
        "kulta, koko, pios",
        "gold, whole/entire, handful",
        ""
    ),
    (
        "hope'ita kourallisen.",
        "hopeita kourallisen.",
        "hopea, kourallinen",
        "silver, a handful",
        ""
    ),
    (
        "Laittoi orjat lietsomahan,",
        "Laittoi orjat lietsomaan,",
        "laittaa, orja, lietsoa",
        "to set/send, slave, to fan/blow (bellows)",
        ""
    ),
    (
        "palkkalaiset painamahan.",
        "palkkalaiset painamaan.",
        "palkkalaiset, painaa",
        "hired hands/workers, to press/push",
        ""
    ),
    (
        "Orjat lietsoi löyhytteli,",
        "Orjat lietsoi löyhytteli,",
        "orja, lietsoa, löyhyttää",
        "slave, to fan/blow bellows, to fan/wave",
        ""
    ),
    (
        "hyvin painoi palkkalaiset:",
        "hyvin painoi palkkalaiset:",
        "hyvin, painaa, palkkalaiset",
        "well, to press/pump, hired hands",
        ""
    ),
    (
        "rauta vellinä venyvi,",
        "rauta vellinä venyy,",
        "rauta, velli, venyä",
        "iron, porridge/gruel, to stretch/melt",
        ""
    ),
    (
        "teräs taipui tahtahana,",
        "teräs taipui tahtahana,",
        "teräs, taipua, tahto",
        "steel, to bend/yield, will",
        ""
    ),
    (
        "hopea vetenä välkkyi,",
        "hopea vetenä välkkyi,",
        "hopea, vesi, välkkyä",
        "silver, water, to glitter/shimmer",
        ""
    ),
    (
        "kulta läikkyi lainehena.",
        "kulta läikkyi lainehena.",
        "kulta, läikkyä, laine",
        "gold, to glitter/ripple, wave",
        ""
    ),
    (
        "Siitä seppo Ilmarinen,",
        "Siitä seppä Ilmarinen,",
        "siitä, seppä, Ilmarinen",
        "then/from that, smith, Ilmarinen",
        ""
    ),
    (
        "takoja iän-ikuinen,",
        "takoja iän-ikuinen,",
        "takoja, ikä, ikuinen",
        "forger/smith, age, eternal",
        ""
    ),
    (
        "katsoi alle ahjoksensa,",
        "katsoi alle ahjonsa,",
        "katsoa, alle, ahjo",
        "to look, under/below, forge/furnace",
        ""
    ),
    (
        "lietsimensä liepehelle:",
        "lietsimensä liepehelle:",
        "liesin, lieve",
        "bellows, edge/flap",
        ""
    ),
    (
        "näki miekan syntyväksi,",
        "näki miekan syntyväksi,",
        "nähdä, miekka, syntyä",
        "to see, sword, to be born/emerge",
        ""
    ),
    (
        "pää kullan kuvauvaksi.",
        "pää kullan kuvautuvaksi.",
        "pää, kulta, kuvautua",
        "head/hilt, gold, to take form/appear",
        ""
    ),
    (
        "Otti ainehet tulesta,",
        "Otti ainehet tulesta,",
        "ottaa, aine, tuli",
        "to take, material/substance, fire",
        ""
    ),
    (
        "tempasi hyvät takehet",
        "tempasi hyvät takehet",
        "tempaista, hyvä, taos",
        "to snatch/grab, good, forging/smithwork",
        ""
    ),
    (
        "ahjosta alasimelle,",
        "ahjosta alasimelle,",
        "ahjo, alasin",
        "forge/furnace, anvil",
        ""
    ),
    (
        "vasarille, valkkamille.",
        "vasarille, valkkamille.",
        "vasara, valkama",
        "hammer, (fulling) mill/water wheel",
        ""
    ),
    (
        "Takoi miekan mieltä myöten,",
        "Takoi miekan mieltä myöten,",
        "takoa, miekka, mieli, myöten",
        "to forge/hammer, sword, mind/will, according to",
        ""
    ),
    (
        "kalvan kaikkien parahan,",
        "kalvan kaikkien parasta,",
        "kalpa, kaikki, paras",
        "sword/blade, all/every, best",
        ""
    ),
    (
        "jonka kullalla kuvasi,",
        "jonka kullalla kuvasi,",
        "joka, kulta, kuvata",
        "which, gold, to decorate/depict",
        ""
    ),
    (
        "hopealla huolitteli.",
        "hopealla huolitteli.",
        "hopea, huolitella",
        "silver, to trim/finish carefully",
        ""
    ),
    (
        "Vaka vanha Väinämöinen",
        "Vankka vanha Väinämöinen",
        "vakaa, vanha, Väinämöinen",
        "steady, old, Väinämöinen",
        ""
    ),
    (
        "tuli tuota katsomahan.",
        "tuli tuota katsomaan.",
        "tulla, tuo, katsoa",
        "to come, that, to look/view",
        ""
    ),
    (
        "Sai miekan tuliteräisen",
        "Sai miekan tuliteräisen",
        "saada, miekka, tuliteräinen",
        "to get/receive, sword, fire-steeled/razor-sharp",
        ""
    ),
    (
        "kätehensä oikeahan.",
        "käteensä oikeaan.",
        "käsi, oikea",
        "hand, right",
        ""
    ),
    (
        "Katselevi, kääntelevi;",
        "Katselee, kääntelee;",
        "katsella, käännellä",
        "to examine/look at, to turn/rotate",
        ""
    ),
    (
        "sanan virkkoi, noin nimesi:",
        "sanan sanoi, noin nimesi:",
        "sana, virkkoa, noin, nimetä",
        "word, to say, thus, to name/call",
        ""
    ),
    (
        '"Onko miekka miestä myöten,',
        '"Onko miekka miestä myöten,',
        "olla, miekka, mies, myöten",
        "to be, sword, man, according to/fitting",
        ""
    ),
    (
        "kalpa kantajan mukahan?",
        "kalpa kantajan mukaan?",
        "kalpa, kantaja, mukaan",
        "sword/blade, bearer/carrier, according to/fitting",
        ""
    ),
    (
        "Olipa miekka miestä myöten,",
        "Olipa miekka miestä myöten,",
        "olla, miekka, mies, myöten",
        "to be, sword, man, according to/fitting",
        ""
    ),
    (
        "kalpa kantajan mukahan,",
        "kalpa kantajan mukaan,",
        "kalpa, kantaja, mukaan",
        "sword/blade, bearer/carrier, according to/fitting",
        ""
    ),
    (
        "jonka kuu kärestä paistoi,",
        "jonka kuu kärestä paistoi,",
        "joka, kuu, kärki, paistaa",
        "which, moon, tip/point/edge, to shine",
        ""
    ),
    (
        "päivä paistoi lappeasta,",
        "päivä paistoi lappeasta,",
        "päivä, paistaa, lape",
        "sun/day, to shine, flat of a blade",
        ""
    ),
    (
        "tähet västistä välötti,",
        "tähdet selästä välkkyivät,",
        "tähti, selkä, välkkyä",
        "star, back/spine, to glitter/sparkle",
        ""
    ),
    (
        "hevonen terällä hirnui,",
        "hevonen terällä hirnui,",
        "hevonen, terä, hirnua",
        "horse, blade/edge, to neigh",
        ""
    ),
    (
        "kasi naukui naulan päässä,",
        "kissa naukui naulan päässä,",
        "kissa, naukua, naula, pää",
        "cat, to meow, nail, head/end",
        ""
    ),
    (
        "penu putkessa puhusi.",
        "pentu putkessa puhui.",
        "pentu, putki, puhua",
        "pup/cub, tube/barrel, to speak",
        ""
    ),
    (
        "Sylkytteli miekkoansa",
        "Heilutti miekkaansa",
        "sylkyttää, miekka",
        "to swing/brandish, sword",
        ""
    ),
    (
        "vuoren rautaisen raossa.",
        "vuoren rautaisen raossa.",
        "vuori, rautainen, rako",
        "mountain, iron (adj.), crack/cleft",
        ""
    ),
    (
        "Itse tuon sanoiksi virkki:",
        "Itse tuon sanoiksi sanoi:",
        "itse, tuo, sana, virkkoa",
        "himself, that, word, to say/utter",
        ""
    ),
    (
        '"Jo minä terällä tällä',
        '"Jo minä terällä tällä',
        "jo, minä, terä, tämä",
        "already/now, I, blade/edge, this",
        ""
    ),
    (
        "vaikka vuoret poikki löisin,",
        "vaikka vuoret poikki löisin,",
        "vaikka, vuori, poikki, lyödä",
        "even if, mountain, in two/across, to strike/hit",
        ""
    ),
    (
        "kalliot kaha jakaisin!",
        "kalliot kahteen jakaisin!",
        "kallio, kahteen, jakaa",
        "rock/cliff, in two, to divide/split",
        ""
    ),
    (
        "Itse seppo Ilmarinen",
        "Itse seppä Ilmarinen",
        "itse, seppä, Ilmarinen",
        "himself, smith, Ilmarinen",
        ""
    ),
    (
        "sanan virkkoi, noin nimesi:",
        "sanan sanoi, noin nimesi:",
        "sana, virkkoa, noin, nimetä",
        "word, to say, thus, to name/call",
        ""
    ),
    (
        '"Milläpä minä poloinen,',
        '"Milläpä minä poloinen,',
        "millä, minä, poloinen",
        "with what, I, poor wretch/wretched one",
        ""
    ),
    (
        "millä, tuima, turveleime,",
        "millä, tuima, turvelemme,",
        "millä, tuima, turvella",
        "with what, fierce/sharp, to protect",
        ""
    ),
    (
        "hyöteleime, vyöteleime",
        "hyödymme, vyötämme",
        "hyötyä, vyöttää",
        "to benefit/avail, to belt/gird",
        ""
    ),
    (
        "maan varalle, veen varalle?",
        "maan varalle, veden varalle?",
        "maa, vara, vesi",
        "land/earth, reserve/provision, water",
        ""
    ),
    (
        "Joko luustoihin lueime,",
        "Joko luustoihin laitamme,",
        "jo, luusto, laittaa",
        "already/now, armor/bone-suit, to put/set",
        ""
    ),
    (
        "rautapaitoihin paneime,",
        "rautapaitoihin paneemme,",
        "rauta, paita, panna",
        "iron, shirt/garment, to put/place",
        ""
    ),
    (
        "teräsvöihin telkitäime?",
        "teräsvöihin telkitäämme?",
        "teräs, vyö, telkitä",
        "steel, belt, to bolt/fasten",
        ""
    ),
    (
        "Mies on luustossa lujempi,",
        "Mies on luustossa lujempi,",
        "mies, luusto, luja",
        "man, armor/bone-suit, sturdy/firm",
        ""
    ),
    (
        "rautapaiassa parempi,",
        "rautapaidassa parempi,",
        "rauta, paita, parempi",
        "iron, shirt/garment, better",
        ""
    ),
    (
        "teräsvyössä tenhoisampi.",
        "teräsvyössä tenhoisampi.",
        "teräs, vyö, tenhoisa",
        "steel, belt, more powerful/enchanting",
        ""
    ),
    (
        "Lähteä luku tulevi,",
        "Lähteä luku tulee,",
        "lähteä, luku, tulla",
        "to depart, count/reckoning, to come",
        ""
    ),
    (
        "liitto käyä kerkiävi.",
        "liitto käydä kerkeää.",
        "liitto, käydä, kerkeä",
        "alliance/agreement, to go, ready/quick",
        ""
    ),
    (
        "Yks' on vanha Väinämöinen,",
        "Yksi on vanha Väinämöinen,",
        "yksi, vanha, Väinämöinen",
        "one, old, Väinämöinen",
        ""
    ),
    (
        "toinen seppo Ilmarinen",
        "toinen seppä Ilmarinen",
        "toinen, seppä, Ilmarinen",
        "second/another, smith, Ilmarinen",
        ""
    ),
    (
        "läksivät hevon hakuhun,",
        "läksivät hevosen hakuun,",
        "lähteä, hevonen, haku",
        "to leave/go, horse, search/seeking",
        ""
    ),
    (
        "kuloharjan kuuntelohon,",
        "kuloharjan kuunteluun,",
        "kuloharja, kuunnella",
        "tawny-maned horse, to listen for",
        ""
    ),
    (
        "suvikunnan suitset vyöllä,",
        "kesäajan suitset vyöllä,",
        "suvi, kunta, suitset, vyö",
        "summer, set/gear, bridle/reins, belt",
        ""
    ),
    (
        "varsan valjahat olalla.",
        "varsan valjahat olalla.",
        "varsa, valjaat, olka",
        "foal/colt, harness, shoulder",
        ""
    ),
    (
        "Kahen etsivät hevoista,",
        "Kahdestaan etsivät hevosia,",
        "kaksi, etsiä, hevonen",
        "two, to search/look for, horse",
        ""
    ),
    (
        "päätä puitse katselevat,",
        "päätä puiden katselevat,",
        "pää, puu, katsella",
        "head, tree, to look/observe",
        ""
    ),
    (
        "tarkasti tähystelevät",
        "tarkasti tähystelevät",
        "tarkasti, tähystää",
        "carefully, to peer/watch",
        ""
    ),
    (
        "ympäri salon sinisen:",
        "ympäri salon sinisen:",
        "ympäri, salo, sininen",
        "around, wilderness/forest, blue",
        ""
    ),
    (
        "löytivät hevon lehosta,",
        "löytivät hevosen lehosta,",
        "löytää, hevonen, lehto",
        "to find, horse, grove",
        ""
    ),
    (
        "kuloharjan kuusikosta.",
        "kuloharjan kuusikosta.",
        "kuloharja, kuusikko",
        "tawny-maned horse, spruce forest",
        ""
    ),
    (
        "Vaka vanha Väinämöinen,",
        "Vankka vanha Väinämöinen,",
        "vakaa, vanha, Väinämöinen",
        "steady, old, Väinämöinen",
        ""
    ),
    (
        "toinen seppo Ilmarinen",
        "toinen seppä Ilmarinen",
        "toinen, seppä, Ilmarinen",
        "second/another, smith, Ilmarinen",
        ""
    ),
    (
        "painoi päähän kullan päitset,",
        "painoi päähän kullan päitset,",
        "painaa, pää, kulta, päitset",
        "to press/put, head, gold, headstall/bridle",
        ""
    ),
    (
        "suvikunnan suitset suuhun.",
        "kesäajan suitset suuhun.",
        "suvi, kunta, suitset, suu",
        "summer, set/gear, reins, mouth",
        ""
    ),
    (
        "Ajoa ratustelevat",
        "Ajoa ratustelevat",
        "ajaa, ratustella",
        "to drive/ride, to stroll/travel",
        ""
    ),
    (
        "kahen miehen rantamaata:",
        "kahden miehen rantamaata:",
        "kaksi, mies, ranta, maa",
        "two, man, shore, land",
        ""
    ),
    (
        "kuului rannalta kujerrus,",
        "kuului rannalta kujerrus,",
        "kuulua, ranta, kujerrus",
        "to be heard, shore, warbling/sobbing sound",
        ""
    ),
    (
        "valitanta valkamalta.",
        "valitanta valkamalta.",
        "valitanta, valkama",
        "lamentation/wailing, harbour/landing place",
        ""
    ),
    (
        "Vaka vanha Väinämöinen",
        "Vankka vanha Väinämöinen",
        "vakaa, vanha, Väinämöinen",
        "steady, old, Väinämöinen",
        ""
    ),
    (
        "sanan virkkoi, noin nimesi:",
        "sanan sanoi, noin nimesi:",
        "sana, virkkoa, noin, nimetä",
        "word, to say, thus, to name/call",
        ""
    ),
    (
        "Siell' on impi itkemässä,",
        "Siellä on impi itkemässä,",
        "siellä, impi, itkeä",
        "there, maiden/girl, to cry/weep",
        ""
    ),
    (
        "kana kaikerrehtamassa!",
        "kana kaikerrehtamassa!",
        "kana, kaikerrehtaa",
        "hen/bird, to cluck/fuss/wail",
        ""
    ),
    (
        "Joko käymme katsomahan,",
        "Joko käymme katsomaan,",
        "jo, käydä, katsoa",
        "now/already, to go, to see/look",
        ""
    ),
    (
        "likeltä tähystämähän?",
        "läheltä tähystämään?",
        "läheltä, tähystää",
        "from nearby/closely, to peer/observe",
        ""
    ),
    (
        "Itse astuvi likemmä,",
        "Itse astuu lähemmäksi,",
        "itse, astua, lähemmäksi",
        "himself, to step/walk, closer",
        ""
    ),
    (
        "meni luota katsomahan.",
        "meni läheltä katsomaan.",
        "mennä, luota, katsoa",
        "to go, from nearby, to look",
        ""
    ),
    (
        "Eipä impi itkekänä",
        "Eipä impi itkekänä",
        "ei, impi, itkeä",
        "not, maiden, to cry/weep",
        ""
    ),
    (
        "eikä kaikerra kananen:",
        "eikä kaikerra kananen:",
        "eikä, kaikerrella, kana",
        "nor, to cluck/wail, hen/bird",
        ""
    ),
    (
        "oli pursi itkemässä,",
        "oli pursi itkemässä,",
        "olla, pursi, itkeä",
        "to be, boat, to cry/weep",
        ""
    ),
    (
        "venonen valittamassa.",
        "venonen valittamassa.",
        "vene, valittaa",
        "boat, to lament/complain",
        ""
    ),
    (
        "Virkki vanha Väinämöinen",
        "Sanoi vanha Väinämöinen",
        "virkkoa, vanha, Väinämöinen",
        "to say/utter, old, Väinämöinen",
        ""
    ),
    (
        "luoksi purren päästyänsä:",
        "luokse purren päästyään:",
        "luokse, pursi, päästä",
        "to/beside, boat, to reach/arrive",
        ""
    ),
    (
        '"Mitä itket, puinen pursi,',
        '"Mitä itket, puinen pursi,',
        "mitä, itkeä, puinen, pursi",
        "what, to cry/weep, wooden, boat",
        ""
    ),
    (
        "vene hankava, valitat?",
        "vene hankava, valitat?",
        "vene, hankava, valittaa",
        "boat, oar-ribbed/rowlocked, to lament",
        ""
    ),
    (
        "Itketkö sä puisuuttasi,",
        "Itketkö sinä puisuuttasi,",
        "itkeä, sinä, puisuus",
        "to cry, you, woodenness/being of wood",
        ""
    ),
    (
        "hankavuuttasi haveksit?",
        "hankavuuttasi haveksit?",
        "hankavuus, haveksia",
        "oar-rib quality, to grieve/regret",
        ""
    ),
    (
        "Pursi puinen vastoavi,",
        "Pursi puinen vastaa,",
        "pursi, puinen, vastata",
        "boat, wooden, to answer/reply",
        ""
    ),
    (
        "vene hankava sanovi:",
        "vene hankava sanoo:",
        "vene, hankava, sanoa",
        "boat, oar-ribbed, to say",
        ""
    ),
    (
        '"Vesille venosen mieli',
        '"Vesille venosen mieli',
        "vesi, vene, mieli",
        "water, boat, mind/wish/desire",
        "The speaking boat expressing longing is a trickster-adjacent motif: animate objects that desire, deceive expectations, and speak the unexpected."
    ),
    (
        "tervaisiltaki teloilta,",
        "tervaisiltaki teloilta,",
        "tervainen, tela",
        "tarred, roller/log (for launching boats)",
        ""
    ),
    (
        "mieli neien miehelähän",
        "mieli neion miehelähän",
        "mieli, neito, mies",
        "mind/wish, maiden, man/husband",
        "Comparing a boat's desire to a maiden's longing for a husband — a playful, subversive metaphor."
    ),
    (
        "korkeastaki ko'ista.",
        "korkeastakin kodista.",
        "korkea, koti",
        "high/tall, home",
        ""
    ),
    (
        "Sitä itken, pursi raukka,",
        "Sitä itken, pursi raukka,",
        "se, itkeä, pursi, raukka",
        "that, to cry, boat, poor wretch",
        "The boat's self-pity is a trickster element: the weak/marginalised voice complaining about being overlooked."
    ),
    (
        "vene vaivainen, valitan:",
        "vene vaivainen, valitan:",
        "vene, vaivainen, valittaa",
        "boat, wretched/poor, to lament",
        ""
    ),
    (
        "itken viejäistä vesille,",
        "itken viejää vesille,",
        "itkeä, viejä, vesi",
        "to cry, one who takes/leads, water",
        ""
    ),
    (
        "laskijaista lainehille.",
        "laskijaa lainehille.",
        "laskija, laine",
        "one who launches/lowers, wave",
        ""
    ),
    (
        '"Sanottihin tehtäessä,',
        '"Sanottiin tehtäessä,',
        "sanoa, tehdä",
        "to say, to make/build",
        "Broken promises and unrealised boasts — a classic trickster grievance; the boat was promised glory but received none."
    ),
    (
        "laulettihin laitettaissa",
        "laulettiin laitettaessa",
        "laulaa, laittaa",
        "to sing, to make/prepare",
        ""
    ),
    (
        "saatavan sotivenettä,",
        "saatavan sotavenettä,",
        "saada, sota, vene",
        "to get/become, war, boat",
        ""
    ),
    (
        "vainopurtta puuhattavan,",
        "vainopurtta puuhattavan,",
        "vaino, pursi, puuhata",
        "pursuit/war, boat, to busy oneself/build",
        ""
    ),
    (
        "tuovan täyteni eloa,",
        "tuovan täyteen eloa,",
        "tuoda, täysi, elo",
        "to bring, full, life/livelihood/grain",
        ""
    ),
    (
        "alustani aartehia:",
        "alustani aarteita:",
        "alusta, aarre",
        "hold/bottom (of boat), treasure",
        ""
    ),
    (
        "ei ole sotahan saatu,",
        "ei ole sotaan saatu,",
        "ei, sota, saada",
        "not, war, to be taken/used",
        ""
    ),
    (
        "eloteillen ensinkänä!",
        "eloteillekaan ensinkään!",
        "elo, tie, ensinkään",
        "livelihood/life, road/path, not at all",
        ""
    ),
    (
        '"Muut purret, pahatki purret,',
        '"Muut purret, pahatki purret,',
        "muu, pursi, paha",
        "other, boat, bad/evil",
        "The boat compares itself enviously to lesser vessels who prosper — trickster's awareness of injustice and social inversion."
    ),
    (
        "ne aina sotia käyvät,",
        "ne aina sotia käyvät,",
        "ne, aina, sota, käydä",
        "they, always, war, to go/engage in",
        ""
    ),
    (
        "tappeloita tallustavat;",
        "tappeloita tallustavat;",
        "tappelu, tallustaa",
        "fight/brawl, to tramp/tread through",
        ""
    ),
    (
        "kolme kertoa kesässä",
        "kolme kertaa kesässä",
        "kolme, kerta, kesä",
        "three, time, summer",
        ""
    ),
    (
        "tuovat täytensä rahoja,",
        "tuovat täytensä rahoja,",
        "tuoda, täysi, raha",
        "to bring, full, money",
        ""
    ),
    (
        "alustansa aartehia.",
        "alustansa aarteita.",
        "alusta, aarre",
        "hold/bottom (of boat), treasure",
        ""
    ),
    (
        "Minä, veistämä venonen,",
        "Minä, veistetty venonen,",
        "minä, veistää, vene",
        "I, to carve/shape, boat",
        ""
    ),
    (
        "satalauta laaittama,",
        "satalaudoin laadittu,",
        "sata, lauta, laittaa",
        "hundred, plank, to make/build",
        ""
    ),
    (
        "tässä lahon lastuillani,",
        "tässä lahon lastuillani,",
        "tässä, laho, lastu",
        "here, rotten, chip/shaving",
        ""
    ),
    (
        "venyn veistännäisilläni.",
        "venyn veistämisessäni.",
        "venyä, veistäminen",
        "to stretch/lie idle, carving/shaping",
        ""
    ),
    (
        "Pahimmatki maan matoset",
        "Pahimmatkin maan madot",
        "paha, maa, mato",
        "worst, land/earth, worm/snake",
        ""
    ),
    (
        "alla kaarien asuvat,",
        "alla kaarien asuvat,",
        "alla, kaari, asua",
        "under, rib (of boat), to live/dwell",
        ""
    ),
    (
        "linnut ilman ilke'immät",
        "linnut ilman ilkeimmät",
        "lintu, ilma, ilkeä",
        "bird, air, nasty/evil",
        ""
    ),
    (
        "pesän pielessä pitävät,",
        "pesän pielessä pitävät,",
        "pesä, pieli, pitää",
        "nest, side/edge, to keep/hold",
        ""
    ),
    (
        "kaikki korven konnikatki",
        "kaikki korven konnatkin",
        "kaikki, korpi, konna",
        "all, wilderness/swamp, rogue/knave/frog",
        ""
    ),
    (
        "kokillani koksentavat.",
        "kokassani kokentavat.",
        "kokka, kokentaa",
        "bow (of boat), to nest/lodge",
        ""
    ),
    (
        "Oisi kahta kaunihimpi,",
        "Olisi kahta kauniimpaa,",
        "olla, kaksi, kaunis",
        "to be, two, beautiful",
        ""
    ),
    (
        "kahta, kolmea parempi",
        "kahta, kolmea parempi",
        "kaksi, kolme, parempi",
        "two, three, better",
        ""
    ),
    (
        "olla mäntynä mäellä,",
        "olla mäntynä mäellä,",
        "olla, mänty, mäki",
        "to be, pine tree, hill",
        ""
    ),
    (
        "petäjänä kankahalla,",
        "petäjänä kankaalla,",
        "petäjä, kangas",
        "pine tree, heath/plain",
        ""
    ),
    (
        "oksilla oravan juosta,",
        "oksilla oravan juosta,",
        "oksa, orava, juosta",
        "branch, squirrel, to run",
        ""
    ),
    (
        "penun alla pyörähellä.",
        "pennun alla pyörähellä.",
        "pentu, alla, pyörähellä",
        "pup/cub, under/below, rolling/frolicking",
        ""
    ),
    (
        "Vaka vanha Väinämöinen",
        "Vankka vanha Väinämöinen",
        "vakaa, vanha, Väinämöinen",
        "steady, old, Väinämöinen",
        ""
    ),
    (
        "tuossa tuon sanoiksi virkki:",
        "tuossa tuon sanoiksi sanoi:",
        "tuossa, tuo, sana, virkkoa",
        "there/then, that, word, to say",
        ""
    ),
    (
        '"Elä itke, puinen pursi,',
        '"Älä itke, puinen pursi,',
        "älä, itkeä, puinen, pursi",
        "do not, to cry/weep, wooden, boat",
        ""
    ),
    (
        "vene hankava, havise!",
        "vene hankava, havise!",
        "vene, hankava, havista",
        "boat, oar-ribbed, to rustle/fuss",
        ""
    ),
    (
        "Kohta saat sotia käyä,",
        "Kohta saat sotia käydä,",
        "kohta, saada, sota, käydä",
        "soon/shortly, to get, war, to go",
        ""
    ),
    (
        "tappeloita tallustella.",
        "tappeloita tallustella.",
        "tappelu, tallustella",
        "fight/brawl, to tramp/wade through",
        ""
    ),
    (
        '"Lienet pursi Luojan luoma,',
        '"Olet pursi Luojan luoma,',
        "olla, pursi, Luoja, luoda",
        "to be, boat, Creator/God, to create",
        "Invoking divine origin to empower the boat — Väinämöinen uses magic speech (a trickster tool) rather than physical force."
    ),
    (
        "Luojan luoma, tuojan tuoma,",
        "Luojan luoma, tuojan tuoma,",
        "Luoja, luoda, tuoja, tuoda",
        "Creator, to create, bringer, to bring",
        ""
    ),
    (
        "syrjin syökseite vetehen,",
        "syrjin syöksikää veteen,",
        "syrjin, syöksiä, vesi",
        "sideways/on its side, to dash/launch, water",
        ""
    ),
    (
        "laion aalloillen ajaite,",
        "laajoille aalloille ajakaa,",
        "laaja, aalto, ajaa",
        "wide, wave, to drive/sail",
        ""
    ),
    (
        "ilman kouran koskematta,",
        "ilman kouran koskematta,",
        "ilman, koura, koskea",
        "without, fist/hand, to touch",
        "The boat launches itself without being touched — magic by words alone; Väinämöinen's power as a word-trickster/shaman."
    ),
    (
        "käen päälle käyttämättä,",
        "käden päälle käyttämättä,",
        "käsi, päälle, käyttää",
        "hand, on top of/upon, to use/apply",
        ""
    ),
    (
        "olkapään ojentamatta,",
        "olkapään ojentamatta,",
        "olkapää, ojentaa",
        "shoulder, to stretch/extend",
        ""
    ),
    (
        "käsivarren vaalimatta!",
        "käsivarren vaalimatta!",
        "käsivarsi, vaalimatta",
        "arm, without tending/lifting",
        ""
    ),
    (
        "Pursi puinen vastoavi,",
        "Pursi puinen vastaa,",
        "pursi, puinen, vastata",
        "boat, wooden, to answer/reply",
        ""
    ),
    (
        "vene hankava sanovi:",
        "vene hankava sanoo:",
        "vene, hankava, sanoa",
        "boat, oar-ribbed, to say",
        ""
    ),
    (
        '"Eipä muu sukuni suuri',
        '"Eipä muu sukuni suuri',
        "ei, muu, suku, suuri",
        "not, other, kin/family, great/large",
        ""
    ),
    (
        "eikä veljeni, venoset,",
        "eikä veljeni, veneet,",
        "eikä, veli, vene",
        "nor, brother, boat",
        ""
    ),
    (
        "lähe työnnyttä vesille,",
        "lähde työnnetyksi vesille,",
        "lähteä, työntää, vesi",
        "to go/launch, to push, water",
        ""
    ),
    (
        "laskematta lainehille,",
        "laskematta lainehille,",
        "laskea, laine",
        "to lower/launch, wave",
        ""
    ),
    (
        "kun ei kourin koskettane,",
        "kun ei kourin kosketa,",
        "kun, ei, koura, koskea",
        "if/when, not, fist/grip, to touch",
        ""
    ),
    (
        "käsivarsin käännettäne.",
        "käsivarsin käännetä.",
        "käsivarsi, kääntää",
        "arm, to turn/manoeuvre",
        ""
    ),
    (
        "Sanoi vanha Väinämöinen:",
        "Sanoi vanha Väinämöinen:",
        "sanoa, vanha, Väinämöinen",
        "to say, old, Väinämöinen",
        ""
    ),
    (
        '"Jos ma sun vesille työnnän,',
        '"Jos ma sinun vesille työnnän,',
        "jos, minä, sinä, vesi, työntää",
        "if, I, you, water, to push/launch",
        ""
    ),
    (
        "joko juokset soutamatta,",
        "joko juokset soutamatta,",
        "jo, juosta, soutaa",
        "now/already, to run, to row",
        ""
    ),
    (
        "airoilla avittamatta,",
        "airoilla avittamatta,",
        "airo, avittaa",
        "oar, to help/assist",
        ""
    ),
    (
        "huoparilla huopimatta,",
        "huoparilla huopimatta,",
        "huopari, huopia",
        "steering oar/rudder, to steer",
        ""
    ),
    (
        "puhumatta purjehesen?",
        "puhumatta purjeisiin?",
        "puhua, purje",
        "to speak/blow, sail",
        ""
    ),
    (
        "Pursi puinen vastoavi,",
        "Pursi puinen vastaa,",
        "pursi, puinen, vastata",
        "boat, wooden, to answer/reply",
        ""
    ),
    (
        "vene hankava sanovi:",
        "vene hankava sanoo:",
        "vene, hankava, sanoa",
        "boat, oar-ribbed, to say",
        ""
    ),
    (
        '"Eipä muu sukuni muuki,',
        '"Eipä muu sukuni muukaan,',
        "ei, muu, suku",
        "not, other, kin/family",
        ""
    ),
    (
        "eikä toinen joukkioni",
        "eikä toinen joukkioni",
        "eikä, toinen, joukko",
        "nor, second/another, group/company",
        ""
    ),
    (
        "juokse sormin soutamatta,",
        "juokse sormein soutamatta,",
        "juosta, sormi, soutaa",
        "to run, finger, to row",
        ""
    ),
    (
        "airoilla avittamatta,",
        "airoilla avittamatta,",
        "airo, avittaa",
        "oar, to help/assist",
        ""
    ),
    (
        "huoparilla huopimatta,",
        "huoparilla huopimatta,",
        "huopari, huopia",
        "steering oar/rudder, to steer",
        ""
    ),
    (
        "puhumatta purjehesen.",
        "puhumatta purjeisiin.",
        "puhua, purje",
        "to speak/blow, sail",
        ""
    ),
    (
        "Vaka vanha Väinämöinen",
        "Vankka vanha Väinämöinen",
        "vakaa, vanha, Väinämöinen",
        "steady, old, Väinämöinen",
        ""
    ),
    (
        "tuosta tuon sanoiksi virkki:",
        "tuosta tuon sanoiksi sanoi:",
        "tuosta, tuo, sana, virkkoa",
        "from that/then, that, word, to say",
        ""
    ),
    (
        '"Joko juokset soutamalla,',
        '"Joko juokset soutamalla,',
        "jo, juosta, soutaa",
        "now/already, to run, to row",
        ""
    ),
    (
        "airoilla avittamalla,",
        "airoilla avittamalla,",
        "airo, avittaa",
        "oar, to help/assist",
        ""
    ),
    (
        "huoparilla huopimalla,",
        "huoparilla huopimalla,",
        "huopari, huopia",
        "steering oar/rudder, to steer",
        ""
    ),
    (
        "puhumalla purjehesen?",
        "puhumalla purjeisiin?",
        "puhua, purje",
        "to speak/blow, sail",
        ""
    ),
    (
        "Pursi puinen vastoavi,",
        "Pursi puinen vastaa,",
        "pursi, puinen, vastata",
        "boat, wooden, to answer/reply",
        ""
    ),
    (
        "vene hankava sanovi:",
        "vene hankava sanoo:",
        "vene, hankava, sanoa",
        "boat, oar-ribbed, to say",
        ""
    ),
    (
        '"Jo vainen sukuni muuki,',
        '"Jo kyllä sukuni muukin,',
        "jo, kyllä, suku, muu",
        "already/yes, yes, kin/family, other",
        ""
    ),
    (
        "kaikki veljeni, venoset,",
        "kaikki veljeni, veneet,",
        "kaikki, veli, vene",
        "all, brother, boat",
        ""
    ),
    (
        "juoksi sormin soutamalla,",
        "juoksi sormein soutamalla,",
        "juosta, sormi, soutaa",
        "to run, finger, to row",
        ""
    ),
    (
        "airoilla avittamalla,",
        "airoilla avittamalla,",
        "airo, avittaa",
        "oar, to help/assist",
        ""
    ),
    (
        "huoparilla huopimalla,",
        "huoparilla huopimalla,",
        "huopari, huopia",
        "steering oar/rudder, to steer",
        ""
    ),
    (
        "puhumalla purjehesen.",
        "puhumalla purjeisiin.",
        "puhua, purje",
        "to speak/blow, sail",
        ""
    ),
    (
        "Siitä vanha Väinämöinen",
        "Siitä vanha Väinämöinen",
        "siitä, vanha, Väinämöinen",
        "then/from that, old, Väinämöinen",
        ""
    ),
    (
        "heitti hiekalle hevosen,",
        "heitti hiekalle hevosen,",
        "heittää, hiekka, hevonen",
        "to throw/leave, sand, horse",
        ""
    ),
    (
        "painoi puuhun marhaminnan,",
        "painoi puuhun marhaminnan,",
        "painaa, puu, marhaminta",
        "to press/tie, tree, tether/rope",
        ""
    ),
    (
        "ohjat oksalle ojenti,",
        "ohjat oksalle ojenti,",
        "ohja, oksa, ojentaa",
        "rein, branch, to extend/hand",
        ""
    ),
    (
        "lykkäsi venon vesille,",
        "lykkäsi veneen vesille,",
        "lykätä, vene, vesi",
        "to push/launch, boat, water",
        ""
    ),
    (
        "lauloi purren lainehille.",
        "lauloi purren lainehille.",
        "laulaa, pursi, laine",
        "to sing, boat, wave",
        "Väinämöinen launches the boat by singing — the paradigmatic trickster-shaman act of reshaping reality through song/word."
    ),
    (
        "Kysytteli puista purtta,",
        "Kysytteli puista purtta,",
        "kysellä, puinen, pursi",
        "to ask/enquire, wooden, boat",
        ""
    ),
    (
        "sanan virkkoi, noin nimesi:",
        "sanan sanoi, noin nimesi:",
        "sana, virkkoa, noin, nimetä",
        "word, to say, thus, to name/call",
        ""
    ),
    (
        '"Oi sie kaareva venonen,',
        '"Oi sinä kaareva venonen,',
        "oi, sinä, kaareva, vene",
        "oh, you, curved/arching, boat",
        ""
    ),
    (
        "pursi puinen, hankaniekka!",
        "pursi puinen, hankaniekka!",
        "pursi, puinen, hankaniekka",
        "boat, wooden, oar-master/expert rower",
        ""
    ),
    (
        "Ootko kaunis kannannalta,",
        "Oletko kaunis kannannalta,",
        "olla, kaunis, kantaa",
        "to be, beautiful, to carry/bear",
        ""
    ),
    (
        "kuin oot kaunis katsonnalta?",
        "kuin olet kaunis katsonnalta?",
        "kuin, olla, kaunis, katsoa",
        "as/how, to be, beautiful, to look",
        ""
    ),
    (
        "Pursi puinen vastoavi,",
        "Pursi puinen vastaa,",
        "pursi, puinen, vastata",
        "boat, wooden, to answer/reply",
        ""
    ),
    (
        "vene hankava sanovi:",
        "vene hankava sanoo:",
        "vene, hankava, sanoa",
        "boat, oar-ribbed, to say",
        ""
    ),
    (
        '"Oonpa kaunis kannannalta',
        '"Olenpa kaunis kannannalta',
        "olla, kaunis, kantaa",
        "to be, beautiful, to carry/bear",
        ""
    ),
    (
        "sekä pohjalta sijava:",
        "sekä pohjalta sijava:",
        "sekä, pohja, sija",
        "and/both, bottom, place/room",
        ""
    ),
    (
        "soutoa sa'an urohon,",
        "soutoa saan urohon,",
        "souto, saada, uros",
        "rowing, to fit/hold, hero/man",
        ""
    ),
    (
        "ilman istua tuhannen.",
        "ilman istua tuhannen.",
        "ilman, istua, tuhat",
        "without/besides, to sit, thousand",
        ""
    ),
    (
        "Siitä vanha Väinämöinen",
        "Siitä vanha Väinämöinen",
        "siitä, vanha, Väinämöinen",
        "then/from that, old, Väinämöinen",
        ""
    ),
    (
        "lauloa hyrähtelevi.",
        "laulaa hyrähtelee.",
        "laulaa, hyrähtelee",
        "to sing, to hum/start singing",
        "Väinämöinen's song as a means of peopling the boat magically — creation by voice is a hallmark trickster-shaman power."
    ),
    (
        "Lauloi ensin laitapuolen",
        "Lauloi ensin laitapuolen",
        "laulaa, ensin, laitapuoli",
        "to sing, first, side (of boat)",
        ""
    ),
    (
        "sukapäitä sulhosia,",
        "sukapäitä sulhosia,",
        "sukapää, sulhonen",
        "stocking-capped (men), young bridegroom/suitor",
        ""
    ),
    (
        "sukapäitä, piipioja,",
        "sukapäitä, piipioja,",
        "sukapää, piippo",
        "stocking-capped (men), young lad/stripling",
        ""
    ),
    (
        "saapasjalkoja jaloja.",
        "saapasjalkoja jaloja.",
        "saapasjalka, jalo",
        "booted-legged (men), noble/fine",
        ""
    ),
    (
        "Lauloi toisen laitapuolen",
        "Lauloi toisen laitapuolen",
        "laulaa, toinen, laitapuoli",
        "to sing, second/other, side (of boat)",
        ""
    ),
    (
        "tinapäitä tyttäriä,",
        "tinapäitä tyttäriä,",
        "tinapää, tytär",
        "tin-/silver-headed (women), daughter",
        ""
    ),
    (
        "tinapäitä, vaskivöitä,",
        "tinapäitä, vaskivöitä,",
        "tinapää, vaskivyö",
        "tin-/silver-headed, copper-belted (women)",
        ""
    ),
    (
        "kultasormia somia.",
        "kultasormia somia.",
        "kultasormi, soma",
        "gold-fingered (women), pretty/lovely",
        ""
    ),
    (
        "Lauloi vielä Väinämöinen",
        "Lauloi vielä Väinämöinen",
        "laulaa, vielä, Väinämöinen",
        "to sing, still/also, Väinämöinen",
        ""
    ),
    (
        "teljot täytehen väkeä,",
        "teljot täyteen väkeä,",
        "teljo, täyteen, väki",
        "thwart/bench (of boat), full, people/crew",
        ""
    ),
    (
        "ne on vanhoa väkeä,",
        "ne on vanhaa väkeä,",
        "ne, vanha, väki",
        "they/those, old, people/folk",
        ""
    ),
    (
        "iän kaiken istunutta,",
        "iän kaiken istunutta,",
        "ikä, kaikki, istua",
        "age, all, to sit",
        ""
    ),
    (
        "kuss' oli vähän sijoa",
        "kussä oli vähän sijaa",
        "kussä, olla, vähän, sija",
        "where, to be, little, room/space",
        ""
    ),
    (
        "nuorukaisilta esinnä.",
        "nuorukaisilta edessä.",
        "nuorukainen, edessä",
        "young man/youth, in front of/before",
        ""
    ),
    (
        "Itse istuvi perähän,",
        "Itse istuu perähän,",
        "itse, istua, perä",
        "himself, to sit, stern/back",
        ""
    ),
    (
        "kokan koivuisen kuvulle,",
        "kokan koivuisen kuvulle,",
        "kokka, koivu, kupu",
        "bow (of boat), birch, dome/curve",
        ""
    ),
    (
        "lasketteli laivoansa.",
        "lasketteli laivaansa.",
        "laskea, laiva",
        "to lower/sail, ship",
        ""
    ),
    (
        "Sanan virkkoi, noin nimesi:",
        "Sanan sanoi, noin nimesi:",
        "sana, virkkoa, noin, nimetä",
        "word, to say, thus, to name/call",
        ""
    ),
    (
        '"Juokse, pursi, puittomia,',
        '"Juokse, pursi, puittomia,',
        "juosta, pursi, puiton",
        "to run, boat, past trees/open",
        ""
    ),
    (
        "vene, väljiä vesiä!",
        "vene, väljiä vesiä!",
        "vene, väljä, vesi",
        "boat, wide/open, water",
        ""
    ),
    (
        "Kule kuplina merellä,",
        "Kule kuplina merellä,",
        "kulkea, kupla, meri",
        "to travel, bubble, sea",
        ""
    ),
    (
        "lumpehina lainehilla!",
        "lumpehina lainehilla!",
        "lumme, laine",
        "water-lily, wave",
        ""
    ),
    (
        "Pani sulhot soutamahan,",
        "Pani sulhot soutamaan,",
        "panna, sulhonen, soutaa",
        "to set/put, suitor/bridegroom, to row",
        ""
    ),
    (
        "neiet ilman istumahan.",
        "neiet ilman istumaan.",
        "neito, ilman, istua",
        "maiden, without, to sit",
        ""
    ),
    (
        "Sulhot souti, airot notkui:",
        "Sulhot souti, airot notkui:",
        "sulhonen, soutaa, airo, notkua",
        "suitor, to row, oar, to bend/flex",
        ""
    ),
    (
        "eipä matka eistykänä.",
        "eipä matka edistynyt.",
        "ei, matka, edistyä",
        "not, journey, to progress/advance",
        ""
    ),
    (
        "Pani neiet soutamahan,",
        "Pani neidot soutamaan,",
        "panna, neito, soutaa",
        "to set, maiden, to row",
        ""
    ),
    (
        "sulhot ilman istumahan.",
        "sulhot ilman istumaan.",
        "sulhonen, ilman, istua",
        "suitor, without, to sit",
        ""
    ),
    (
        "Neiet souti, sormet notkui:",
        "Neidot souti, sormet notkui:",
        "neito, soutaa, sormi, notkua",
        "maiden, to row, finger, to bend",
        ""
    ),
    (
        "eipä matka eistykänä.",
        "eipä matka edistynyt.",
        "ei, matka, edistyä",
        "not, journey, to progress/advance",
        ""
    ),
    (
        "Muutti vanhat soutamahan,",
        "Muutti vanhat soutamaan,",
        "muuttaa, vanha, soutaa",
        "to change/put, old, to row",
        ""
    ),
    (
        "nuoret päältä katsomahan.",
        "nuoret päältä katsomaan.",
        "nuori, päältä, katsoa",
        "young, from above/over, to look",
        ""
    ),
    (
        "Vanhat souti, päät vapisi:",
        "Vanhat souti, päät vapisi:",
        "vanha, soutaa, pää, vapista",
        "old, to row, head, to tremble/shake",
        ""
    ),
    (
        "eipä vielä matka eisty.",
        "eipä vielä matka edistynyt.",
        "ei, vielä, matka, edistyä",
        "not, yet, journey, to progress",
        ""
    ),
    (
        "Siitä seppo Ilmarinen",
        "Siitä seppä Ilmarinen",
        "siitä, seppä, Ilmarinen",
        "then, smith, Ilmarinen",
        ""
    ),
    (
        "itse istui soutamahan:",
        "itse istui soutamaan:",
        "itse, istua, soutaa",
        "himself, to sit, to row",
        ""
    ),
    (
        "jopa juoksi puinen pursi,",
        "jopa juoksi puinen pursi,",
        "jo, juosta, puinen, pursi",
        "now/already, to run, wooden, boat",
        ""
    ),
    (
        "pursi juoksi, matka joutui.",
        "pursi juoksi, matka joutui.",
        "pursi, juosta, matka, joutua",
        "boat, to run, journey, to proceed/hasten",
        ""
    ),
    (
        "Loitos kuului airon loiske,",
        "Loitolle kuului airon loiske,",
        "loitos, kuulua, airo, loiske",
        "far away, to be heard, oar, splash",
        ""
    ),
    (
        "kauas hankojen hamina.",
        "kauas hankojen hamina.",
        "kauas, hanko, hamina",
        "far, rowlock/thole, din/noise",
        ""
    ),
    (
        "Soutavi sorehtelevi:",
        "Soutaa sorehteli:",
        "soutaa, sorehtaa",
        "to row, to row gracefully/powerfully",
        ""
    ),
    (
        "teljot rytkyi, laiat notkui,",
        "teljot rytkyi, laiat notkui,",
        "teljo, rytkkyä, laia, notkua",
        "thwart/bench, to crack/creak, side of boat, to flex",
        ""
    ),
    (
        "airot piukki pihlajaiset,",
        "airot piuksivat pihlajaiset,",
        "airo, piukata, pihlaja",
        "oar, to creak/squeak, rowan (tree)",
        ""
    ),
    (
        "airon pyörät pyinä vinkui,",
        "airon pyörät pyinä vinkui,",
        "airo, pyörä, py, vingata",
        "oar, wheel/circle, grouse, to whine/squeal",
        ""
    ),
    (
        "terät teirinä kukerti,",
        "terät teirinä kukerti,",
        "terä, teiri, kukertaa",
        "blade/edge, black grouse, to cock/crow",
        ""
    ),
    (
        "nenä joikui joutsenena,",
        "nenä joikui joutsenena,",
        "nenä, joikata, joutsen",
        "nose/prow, to yoik/sing, swan",
        ""
    ),
    (
        "perä kaarskui kaarnehena,",
        "perä kaarskui kaarnehena,",
        "perä, kaarskahtaa, kaarnas",
        "stern, to croak, raven/crow",
        ""
    ),
    (
        "hangat hanhina havisi.",
        "hangat hanhina havisi.",
        "hanko, hanhi, havista",
        "rowlock, goose, to rustle/swish",
        ""
    ),
    (
        "Itse vanha Väinämöinen",
        "Itse vanha Väinämöinen",
        "itse, vanha, Väinämöinen",
        "himself, old, Väinämöinen",
        ""
    ),
    (
        "laskea karehtelevi",
        "laskea karehtelee",
        "laskea, karehtelee",
        "to sail/descend, to glide/sail smoothly",
        ""
    ),
    (
        "perässä punaisen purren,",
        "perässä punaisen purren,",
        "perässä, punainen, pursi",
        "at the stern, red, boat",
        ""
    ),
    (
        "melan vartevan varassa.",
        "melan vartevan varassa.",
        "mela, vartevas, varassa",
        "paddle/oar, sturdy/able, relying on",
        ""
    ),
    (
        "Niemi matkalla näkyvi,",
        "Niemi matkalla näkyy,",
        "niemi, matka, näkyä",
        "cape/headland, journey, to be seen",
        ""
    ),
    (
        "kylä kurja kuumottavi.",
        "kylä kurja kuumottaa.",
        "kylä, kurja, kuumottaa",
        "village, miserable/wretched, to shimmer/loom",
        ""
    ),
    (
        "Ahti niemellä asuvi,",
        "Ahti niemellä asuu,",
        "Ahti, niemi, asua",
        "Ahti (= Lemminkäinen), cape, to live",
        ""
    ),
    (
        "Kauko niemen kainalossa.",
        "Kauko niemen kainalossa.",
        "Kauko (= Lemminkäinen), niemi, kainalossa",
        "Kauko (name for Lemminkäinen), cape, in the armpit/nook",
        "Lemminkäinen (Ahti/Kauko) is the quintessential trickster of the Kalevala: capricious, lustful, boastful, cunning."
    ),
    (
        "Kalatuutta Kauko itki,",
        "Kalatuutta Kauko itki,",
        "kala, tuutta, Kauko, itkeä",
        "fish, lack/need, Kauko, to cry/weep",
        "Lemminkäinen crying over lack of fish — trickster's perpetual dissatisfaction and hunger."
    ),
    (
        "leivätyyttä Lemminkäinen,",
        "leivätyyttä Lemminkäinen,",
        "leipä, tyytyminen, Lemminkäinen",
        "bread, lack/need, Lemminkäinen",
        ""
    ),
    (
        "Ahti aitan pieneyttä,",
        "Ahti aitan pieneyttä,",
        "Ahti, aitta, pienuus",
        "Ahti, storehouse/granary, smallness",
        ""
    ),
    (
        "veitikkä osan vähyyttä.",
        "veitikka osan vähyyttä.",
        "veitikka, osa, vähyys",
        "rascal/rogue, share/portion, scarcity",
        "\"Veitikka\" (rascal/rogue) is a direct trickster epithet for Lemminkäinen."
    ),
    (
        "Veisti laitoja venehen,",
        "Veisti laitoja venehen,",
        "veistää, laita, vene",
        "to carve/shape, side (of boat), boat",
        ""
    ),
    (
        "uuen purren pohjapuuta",
        "uuden purren pohjapuuta",
        "uusi, pursi, pohja, puu",
        "new, boat, bottom, wood",
        ""
    ),
    (
        "päässä pitkän nälkäniemen,",
        "päässä pitkän nälkäniemen,",
        "pää, pitkä, nälkä, niemi",
        "head/end, long, hunger, cape/headland",
        ""
    ),
    (
        "paltalla kylän katalan.",
        "paltalla kylän katalan.",
        "palta, kylä, katala",
        "edge/shore, village, miserable/wretched",
        ""
    ),
    (
        "Se oli korvalta korea,",
        "Se oli korvalta korea,",
        "se, korva, korea",
        "it/that, ear/side, handsome/fine",
        ""
    ),
    (
        "silmältä sitäi parempi.",
        "silmältä sitäkin parempi.",
        "silmä, se, parempi",
        "eye, it/that, better",
        ""
    ),
    (
        "Loi silmänsä luotehelle,",
        "Loi silmänsä luoteeseen,",
        "luoda, silmä, luode",
        "to cast, eye, northwest",
        ""
    ),
    (
        "käänti päätä päivän alle:",
        "käänsi päätä päivän alle:",
        "kääntää, pää, päivä, alle",
        "to turn, head, sun/day, under",
        ""
    ),
    (
        "kaaren kaukoa näkevi,",
        "kaaren kaukoa näkee,",
        "kaari, kaukaa, nähdä",
        "arc/bow, from far, to see",
        ""
    ),
    (
        "pilven longan loitompata.",
        "pilven longan loitompaa.",
        "pilvi, lonka, loitompaa",
        "cloud, arm/streak, further away",
        ""
    ),
    (
        "Eipä kaari ollutkana",
        "Eipä kaari ollutkaan",
        "ei, kaari, olla",
        "not, arc/bow, to be",
        ""
    ),
    (
        "eikä pieni pilven lonka:",
        "eikä pieni pilven lonka:",
        "eikä, pieni, pilvi, lonka",
        "nor, small, cloud, arm/streak",
        ""
    ),
    (
        "oli pursi kulkemassa,",
        "oli pursi kulkemassa,",
        "olla, pursi, kulkea",
        "to be, boat, to travel",
        ""
    ),
    (
        "venonen vaeltamassa",
        "venonen vaeltamassa",
        "vene, vaeltaa",
        "boat, to wander/travel",
        ""
    ),
    (
        "selvällä meren selällä,",
        "selvällä meren selällä,",
        "selvä, meri, selkä",
        "clear/open, sea, open water",
        ""
    ),
    (
        "ulapalla aukealla;",
        "ulapalla aukealla;",
        "ulappa, avoin",
        "open sea/expanse, open",
        ""
    ),
    (
        "mies puhas perässä purren,",
        "mies puhdas perässä purren,",
        "mies, puhdas, perässä, pursi",
        "man, clean/pure, at the stern, boat",
        ""
    ),
    (
        "mies sorea soutimilla.",
        "mies sorea soutimilla.",
        "mies, sorea, soudin",
        "man, handsome/graceful, oar",
        ""
    ),
    (
        "Sanoi lieto Lemminkäinen:",
        "Sanoi lieto Lemminkäinen:",
        "sanoa, lieto, Lemminkäinen",
        "to say, carefree/fickle, Lemminkäinen",
        "\"Lieto\" (carefree/fickle/reckless) is Lemminkäinen's standing epithet — the trickster defined by irresponsibility."
    ),
    (
        '"En mä tunne tuota purtta,',
        '"En minä tunne tuota purtta,',
        "ei, minä, tuntea, tuo, pursi",
        "not, I, to know/recognize, that, boat",
        ""
    ),
    (
        "keksi kelvoista venettä;",
        "keksi kelvoista venettä;",
        "keksiä, kelvoinen, vene",
        "to make out/spot, worthy/good, boat",
        ""
    ),
    (
        "souten Suomesta tulevi,",
        "souten Suomesta tulee,",
        "soutaa, Suomi, tulla",
        "to row, Finland, to come",
        ""
    ),
    (
        "airon iske'in iästä,",
        "airon iskuin idästä,",
        "airo, isku, itä",
        "oar, stroke/blow, east",
        ""
    ),
    (
        "melan luoen luotehesen.",
        "melan luoden luoteeseen.",
        "mela, luoda, luode",
        "paddle, to cast/strike, northwest",
        ""
    ),
    (
        "Jo huhuta huikahutti,",
        "Jo huhuta huikahutti,",
        "jo, huhuilla, huikahtaa",
        "already/now, to hoot/call, to shout out",
        ""
    ),
    (
        "mäjellytti, mäikähytti,",
        "mäjähytti, mäikähytti,",
        "mäjähtää, mäikähtää",
        "to shout/boom, to cry out",
        ""
    ),
    (
        "huuti mies nenästä niemen,",
        "huusi mies nenästä niemen,",
        "huutaa, mies, nenä, niemi",
        "to shout, man, tip/end, cape",
        "Lemminkäinen's brazen shouting across the water — trickster's lack of restraint and love of noise/attention."
    ),
    (
        "verevä vesien poikki:",
        "verevä vesien poikki:",
        "verevä, vesi, poikki",
        "vigorous/ruddy, water, across",
        ""
    ),
    (
        '"Kenen on veno vesillä,',
        '"Kenen on vene vesillä,',
        "kenen, vene, vesi",
        "whose, boat, water",
        ""
    ),
    (
        "kenen laiva lainehilla?",
        "kenen laiva lainehilla?",
        "kenen, laiva, laine",
        "whose, ship, wave",
        ""
    ),
    (
        "Miehet purresta puhuvat",
        "Miehet purresta puhuvat",
        "mies, pursi, puhua",
        "man, boat, to speak",
        ""
    ),
    (
        "sekä vaimot vastoavat:",
        "sekä vaimot vastaavat:",
        "sekä, vaimo, vastata",
        "and/both, woman/wife, to answer",
        ""
    ),
    (
        '"Mi olet mies metsän asuja,',
        '"Mikä olet mies metsän asuja,',
        "mikä, olla, mies, metsä, asuja",
        "what/who, to be, man, forest, dweller",
        ""
    ),
    (
        "uros korven kolkuttaja,",
        "uros korven kolkuttaja,",
        "uros, korpi, kolkuttaa",
        "hero/man, wilderness/swamp, to knock/wander",
        ""
    ),
    (
        "kun et tunne tuota purtta,",
        "kun et tunne tuota purtta,",
        "kun, ei, tuntea, tuo, pursi",
        "when/if, not, to know, that, boat",
        ""
    ),
    (
        "keksi Väinölän venettä,",
        "keksi Väinölän venettä,",
        "keksiä, Väinölä, vene",
        "to recognize, Väinölä (Väinämöinen's home), boat",
        ""
    ),
    (
        "et tunne peräurosta",
        "et tunne peräurosta",
        "ei, tuntea, perä, uros",
        "not, to know, stern, hero/man",
        ""
    ),
    (
        "etkä miestä airollista?",
        "etkä miestä airollista?",
        "eikä, mies, airo",
        "nor, man, oar",
        ""
    ),
    (
        "Sanoi lieto Lemminkäinen:",
        "Sanoi lieto Lemminkäinen:",
        "sanoa, lieto, Lemminkäinen",
        "to say, carefree/fickle, Lemminkäinen",
        ""
    ),
    (
        '"Jo tunnen peränpitäjän',
        '"Jo tunnen peränpitäjän',
        "jo, tuntea, peränpitäjä",
        "already/now, to know, helmsman/steersman",
        ""
    ),
    (
        "ja älyän airollisen:",
        "ja älyän airollisen:",
        "ja, älytä, airollinen",
        "and, to figure out, oarsman",
        ""
    ),
    (
        "vaka vanha Väinämöinen",
        "vankka vanha Väinämöinen",
        "vakaa, vanha, Väinämöinen",
        "steady, old, Väinämöinen",
        ""
    ),
    (
        "itse on perän piossa,",
        "itse on perässä piossa,",
        "itse, perä, pio",
        "himself, stern, grasp/rudder-handle",
        ""
    ),
    (
        "Ilmarinen airollisna.",
        "Ilmarinen airollisena.",
        "Ilmarinen, airollinen",
        "Ilmarinen, oarsman",
        ""
    ),
    (
        "Minnekkä menette, miehet,",
        "Minne menette, miehet,",
        "minne, mennä, mies",
        "where, to go, man",
        ""
    ),
    (
        "kunne läksitte, urohot?",
        "minne läksitte, urohot?",
        "minne, lähteä, uros",
        "where, to depart, hero/man",
        ""
    ),
    (
        "Sanoi vanha Väinämöinen:",
        "Sanoi vanha Väinämöinen:",
        "sanoa, vanha, Väinämöinen",
        "to say, old, Väinämöinen",
        ""
    ),
    (
        '"Kohti pohjaista kulemme,',
        '"Kohti pohjaista kullemme,',
        "kohti, pohjainen, kulkea",
        "towards, northern, to travel",
        ""
    ),
    (
        "kohti kuohuja kovia,",
        "kohti kuohuja kovia,",
        "kohti, kuohu, kova",
        "towards, rapid/surge, hard/rough",
        ""
    ),
    (
        "lakkipäitä lainehia:",
        "lakkipäitä lainehia:",
        "lakkipää, laine",
        "cap-headed/crested (waves), wave",
        ""
    ),
    (
        "sampoa tapoamahan,",
        "sampoa tavoittamaan,",
        "sampo, tavoittaa",
        "Sampo, to reach/attain",
        ""
    ),
    (
        "kirjokantta katsomahan",
        "kirjokantta katsomaan",
        "kirjokansi, katsoa",
        "decorated lid, to see/view",
        ""
    ),
    (
        "Pohjolan kivimäestä,",
        "Pohjolan kivimäestä,",
        "Pohjola, kivimäki",
        "Pohjola, rocky hill",
        ""
    ),
    (
        "vaaran vaskisen sisästä.",
        "vaaran vaskisen sisästä.",
        "vaara, vaskinen, sisä",
        "hill, copper/bronze (adj.), inside",
        ""
    ),
    (
        "Sanoi lieto Lemminkäinen:",
        "Sanoi lieto Lemminkäinen:",
        "sanoa, lieto, Lemminkäinen",
        "to say, carefree/fickle, Lemminkäinen",
        ""
    ),
    (
        '"Ohoh vanha Väinämöinen!',
        '"Ohoh vanha Väinämöinen!',
        "ohoh, vanha, Väinämöinen",
        "hey/oh, old, Väinämöinen",
        ""
    ),
    (
        "Otapa minua, miestä,",
        "Ota minua, miestä,",
        "ottaa, minä, mies",
        "to take, I/me, man",
        "Lemminkäinen pushes himself uninvited into the quest — typical trickster gate-crashing."
    ),
    (
        "urohoksi kolmanneksi,",
        "urohoksi kolmanneksi,",
        "uros, kolmas",
        "hero/man, third",
        ""
    ),
    (
        "kun saat sammon nostantahan,",
        "kun saat sammon nostamiseen,",
        "kun, saada, sampo, nostaa",
        "when/if, to get, Sampo, to raise/lift",
        ""
    ),
    (
        "kirjokannen kannantahan!",
        "kirjokannen kantamiseen!",
        "kirjokansi, kantaa",
        "decorated lid, to carry",
        ""
    ),
    (
        "Vielä mieki miesnä maksan,",
        "Vielä minäkin miehenä maksan,",
        "vielä, minä, mies, maksaa",
        "still/yet, I, man, to pay/be worth",
        "Lemminkäinen boasting his worth — the trickster's self-promotion."
    ),
    (
        "jos saisi tapella tarve:",
        "jos saisi tapella tarve:",
        "jos, saada, tapella, tarve",
        "if, to get, to fight, need/necessity",
        ""
    ),
    (
        "annan käskyn kämmenille,",
        "annan käskyn kämmenille,",
        "antaa, käsky, kämmen",
        "to give, command/order, palm (of hand)",
        ""
    ),
    (
        "olkapäilleni opaston.",
        "olkapäilleni opaston.",
        "olkapää, opastus",
        "shoulder, guidance/instruction",
        ""
    ),
    (
        "Vaka vanha Väinämöinen",
        "Vankka vanha Väinämöinen",
        "vakaa, vanha, Väinämöinen",
        "steady, old, Väinämöinen",
        ""
    ),
    (
        "otti miehen matkoihinsa,",
        "otti miehen matkoihinsa,",
        "ottaa, mies, matka",
        "to take, man, journey",
        ""
    ),
    (
        "veitikän venosehensa.",
        "veitikän veneesensä.",
        "veitikka, vene",
        "rascal/rogue, boat",
        "Again \"veitikka\" (rascal/rogue) — Lemminkäinen's trickster identity is embedded in his very name here."
    ),
    (
        "Se on lieto Lemminkäinen",
        "Se on lieto Lemminkäinen",
        "se, lieto, Lemminkäinen",
        "he, carefree/fickle, Lemminkäinen",
        ""
    ),
    (
        "jo tulla tuhuttelevi,",
        "jo tulla tuhuttelee,",
        "jo, tulla, tuhutella",
        "already, to come, to bustle/rush in noisily",
        "Lemminkäinen rushes in noisily and clumsily — slapstick trickster arrival."
    ),
    (
        "käyä luikerrehtelevi.",
        "käydä luikerrehtelee.",
        "käydä, luikerrella",
        "to go, to slither/creep/sneak",
        "\"Luikerrella\" (to slither/sneak) is a trickster movement — cunning, serpentine motion."
    ),
    (
        "Tuopi laian tullessansa",
        "Tuo laidan tullessaan",
        "tuoda, laita, tulla",
        "to bring, side (of boat), to come",
        ""
    ),
    (
        "venehesen Väinämöisen.",
        "veneeseen Väinämöisen.",
        "vene, Väinämöinen",
        "boat, Väinämöinen",
        ""
    ),
    (
        "Sanoi vanha Väinämöinen:",
        "Sanoi vanha Väinämöinen:",
        "sanoa, vanha, Väinämöinen",
        "to say, old, Väinämöinen",
        ""
    ),
    (
        '"Oisi puuta purressani,',
        '"Olisi puuta purressani,',
        "olla, puu, pursi",
        "to be, wood, boat",
        ""
    ),
    (
        "laitoa venehessäni,",
        "laitaa veneessäni,",
        "laita, vene",
        "side (of boat), boat",
        ""
    ),
    (
        "parahiksi painoaki.",
        "parhaaksi painoksi.",
        "paras, paino",
        "best, weight/ballast",
        ""
    ),
    (
        "Miksi laitat laitoasi,",
        "Miksi laitat laitaasi,",
        "miksi, laittaa, laita",
        "why, to damage/put, side (of boat)",
        ""
    ),
    (
        "puuta purtehen liseät?",
        "puuta veneeseen lisäät?",
        "puu, vene, lisätä",
        "wood, boat, to add/increase",
        ""
    ),
    (
        "Sanoi lieto Lemminkäinen:",
        "Sanoi lieto Lemminkäinen:",
        "sanoa, lieto, Lemminkäinen",
        "to say, carefree/fickle, Lemminkäinen",
        ""
    ),
    (
        '"Ei vara venettä kaa\'a,',
        '"Ei vara venettä kaada,',
        "ei, vara, vene, kaataa",
        "not, reserve/extra, boat, to capsize/tip",
        "A classic trickster proverb-like rationalization: \"extra provision won't sink the boat\" — Lemminkäinen justifying his intrusion with folk wisdom."
    ),
    (
        "tuki suovoa tuhoa.",
        "tuki suovoa tuhoa.",
        "tuki, suo, tuho",
        "support/prop, swamp/bog, destruction/ruin",
        ""
    ),
    (
        "Use'in merellä Pohjan",
        "Useiten merellä Pohjan",
        "usein, meri, Pohjola",
        "often, sea, Pohjola (the North)",
        ""
    ),
    (
        "tuuli laitoa kysyvi,",
        "tuuli laitaa kysyy,",
        "tuuli, laita, kysyä",
        "wind, side (of boat), to ask/need",
        ""
    ),
    (
        "vastatuuli varppehia.",
        "vastatuuli varppehia.",
        "vastatuuli, varppa",
        "headwind, warp/rope (for mooring)",
        ""
    ),
    (
        "Sanoi vanha Väinämöinen:",
        "Sanoi vanha Väinämöinen:",
        "sanoa, vanha, Väinämöinen",
        "to say, old, Väinämöinen",
        ""
    ),
    (
        '"Sentähen sotavenosen',
        '"Sentähden sotavenosen',
        "sentähden, sota, vene",
        "therefore/that is why, war, boat",
        ""
    ),
    (
        "rinta rautahan rakettu",
        "rinta rautahan rakettu",
        "rinta, rauta, rakentaa",
        "chest/prow, iron, to build/construct",
        ""
    ),
    (
        "ja tehty teräsnenähän,",
        "ja tehty teräsnenähän,",
        "ja, tehdä, teräs, nenä",
        "and, to make, steel, nose/prow",
        ""
    ),
    (
        "jottei tuulen tuiki vieä",
        "jottei tuulen tuiki vieä",
        "jottei, tuuli, tuiki, viedä",
        "so that not, wind, utterly, to carry away",
        ""
    ),
    (
        "eikä viskoa vihurin.",
        "eikä viskoa vihurin.",
        "eikä, viskata, vihuri",
        "nor, to toss/fling, gale",
        ""
    ),
]

# Write rows
for row_idx, (a, b, c, d, e) in enumerate(verses, start=2):
    ws.cell(row=row_idx, column=1, value=a)
    ws.cell(row=row_idx, column=2, value=b)
    ws.cell(row=row_idx, column=3, value=c)
    ws.cell(row=row_idx, column=4, value=d)
    ws.cell(row=row_idx, column=5, value=e)

# Formatting
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

trickster_fill = PatternFill("solid", start_color="FFF3CD")
alt_fill = PatternFill("solid", start_color="F8F9FA")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
    row_num = row[0].row
    has_trickster = bool(row[4].value)
    for cell in row:
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
        if has_trickster:
            cell.fill = trickster_fill
        elif row_num % 2 == 0:
            cell.fill = alt_fill

# Column widths
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 35
ws.column_dimensions["E"].width = 55

# Row heights
for row_num in range(1, ws.max_row + 1):
    ws.row_dimensions[row_num].height = 15 if row_num == 1 else 30

ws.row_dimensions[1].height = 35

# Freeze header
ws.freeze_panes = "A2"

# Legend note
legend_row = ws.max_row + 2
ws.cell(row=legend_row, column=1, value="Väri/Color: Keltainen / Yellow = säe liittyy trickster-teemaan / verse connected to trickster theme")
ws.cell(row=legend_row, column=1).font = Font(name="Arial", size=9, italic=True, color="7B6E00")
ws.cell(row=legend_row, column=1).fill = PatternFill("solid", start_color="FFF3CD")

wb.save("kalevala_runo39.xlsx")
print("Done!")
