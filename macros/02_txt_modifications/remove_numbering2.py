import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

PATTERN = re.compile(r'^\d+\.\s*')

def lines_to_excel(lines, path, header, header_color):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lines"
    ws.column_dimensions['A'].width = 80

    # Header row
    ws['A1'] = header
    ws['A1'].font = Font(name='Arial', bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color=header_color)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # One line per row — '@' number format forces Excel to treat value as plain text,
    # which prevents it from squishing multiple lines into one cell on open.
    for i, line in enumerate(lines, start=2):
        cell = ws.cell(row=i, column=1, value=line.rstrip('\n'))
        cell.font = Font(name='Arial')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        cell.number_format = '@'

    wb.save(path)

def process(input_path, before_path='before.xlsx', after_path='after.xlsx'):
    with open(input_path, 'r', encoding='utf-8') as f:
        original_lines = f.readlines()

    cleaned_lines = [PATTERN.sub('', line) for line in original_lines]

    lines_to_excel(original_lines, before_path, 'BEFORE', '4472C4')  # blue header
    lines_to_excel(cleaned_lines,  after_path,  'AFTER',  '70AD47')  # green header

    changed = sum(1 for a, b in zip(original_lines, cleaned_lines) if a != b)
    print(f"Processed {len(original_lines)} lines — {changed} line(s) modified.")
    print(f"  Before → {before_path}")
    print(f"  After  → {after_path}")

if __name__ == '__main__':
    # if len(sys.argv) < 2:
    #     print("Usage: python remove_numbering.py <input_file> [before.xlsx] [after.xlsx]")
    #     sys.exit(1)

    input_file  = "Mieleni.txt"
    before_path = 'before.xlsx'
    after_path  = 'after.xlsx'

    process(input_file, before_path, after_path)
